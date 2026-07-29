"""
builder.py — Discovers analysis sessions from a root path and builds an HDF5 database.

HDF5 output hierarchy used by the dashboard:

    <root_var>/
      <CohortName>/
        <MouseName>/
          <SessionType>/
            <SessionNumber>/
              CalciumData/
                SpatialFootprints  (N, W, H)  — transposed from (N, H, W) for MATLAB compat
                TemporalFootprints (N, T)
                MaxProjection      (W, H)     — transposed from (H, W) for MATLAB compat
"""

import os
import re
import csv
import glob
from datetime import datetime, timezone
from uuid import uuid4

import h5py
import numpy as np
from typing import Callable, Dict, List, Optional, Tuple, Any

from .loader import get_loader, DataNotFoundError
from ..cell_registration_dashboard.quality import cell_quality_metrics

ANALYSIS_SUFFIXES = [
    "caiman-analysis",
    "minian-analysis",
    "min1pipe",
]

DEFAULT_ANALYSIS_PATTERNS = {
    "caiman-analysis": "caiman-analysis/**/caiman_results.hdf5",
    "minian-analysis": "minian-analysis/**/A.zarr",
    "min1pipe": "**/*_data_processed*.mat",
}

REQUIRED_FIELDS = ["CohortName", "MouseName", "SessionType", "SessionNumber"]
ROOT_VAR = "Database"


# ──────────────────────────────────────────────────────────────────────────────
# Discovery
# ──────────────────────────────────────────────────────────────────────────────

def find_analysis_target(session_dir: str, pattern: str) -> Tuple[str, int]:
    """Return the newest recursive file/folder match inside a session folder."""
    session_dir = os.path.normpath(session_dir)
    pattern = (pattern or "").strip().replace("\\", os.sep).replace("/", os.sep)
    if not pattern:
        raise ValueError("An analysis result pattern is required.")
    if os.path.isabs(pattern) or ".." in pattern.split(os.sep):
        raise ValueError("Analysis result patterns must be relative to the session folder.")

    recursive_pattern = pattern if "**" in pattern else os.path.join("**", pattern)
    candidates = {
        os.path.normpath(path)
        for path in glob.glob(
            os.path.join(session_dir, recursive_pattern),
            recursive=True,
        )
        if os.path.isfile(path) or os.path.isdir(path)
    }
    if not candidates:
        raise DataNotFoundError(
            f"No analysis result matching '{pattern}' found under: {session_dir}"
        )

    def _sort_key(path: str):
        try:
            modified = os.stat(path).st_mtime_ns
        except OSError:
            modified = -1
        return modified, path.casefold()

    return max(candidates, key=_sort_key), len(candidates)


def discover_sessions(
    root_path: str,
    analysis_type: str,
    depth_rules: Optional[List[Dict[str, Any]]] = None,
    analysis_pattern: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    Finds all *analysis_type* folders under *root_path*.

    When *depth_rules* are supplied the search is depth-targeted: we know
    exactly how many directory levels separate the root from the analysis
    folder, so we use os.listdir at each level instead of a full recursive
    os.walk.  This avoids stat-ing thousands of files inside analysis
    folders on network drives.

    When *depth_rules* is None we fall back to the legacy recursive folder
    search. An empty list means the root itself is the session folder.

    Returns
    -------
    List of dicts:
        {
            "rel_parts":    list[str]  path parts relative to root_path
            "analysis_dir": str        absolute path to the selected result
            "session_dir":  str        absolute path to the session folder
        }
    """
    results = []
    root_path = os.path.normpath(root_path)
    target = analysis_type.lower()
    is_min1pipe = target == "min1pipe"

    def _contains_min1pipe_result(directory: str, entries: List[str]) -> bool:
        return any(
            re.search(r"_data_processed(?:_refined)?\.mat$", name, re.IGNORECASE)
            and os.path.isfile(os.path.join(directory, name))
            for name in entries
        )

    # ── Depth-targeted fast path ───────────────────────────────────────────
    if depth_rules is not None:
        # depth_rules already excludes the root folder itself (off-by-one fix),
        # so len(depth_rules) == number of directory levels between root and the
        # session folder that directly contains the -analysis subdirectory.
        expected_depth = len(depth_rules)

        def _descend(current_dir: str, parts: List[str], depth: int):
            if depth == expected_depth:
                if analysis_pattern:
                    try:
                        selected, match_count = find_analysis_target(
                            current_dir, analysis_pattern
                        )
                    except (DataNotFoundError, OSError, ValueError):
                        return
                    results.append({
                        "rel_parts": list(parts),
                        "analysis_dir": selected,
                        "session_dir": current_dir,
                        "analysis_match_count": match_count,
                    })
                    return
                try:
                    entries = os.listdir(current_dir)
                except PermissionError:
                    return
                if is_min1pipe and _contains_min1pipe_result(current_dir, entries):
                    results.append({
                        "rel_parts": list(parts),
                        "analysis_dir": current_dir,
                        "session_dir": current_dir,
                    })
                    return
                for entry in entries:
                    if entry.lower() == target:
                        full = os.path.join(current_dir, entry)
                        if os.path.isdir(full):
                            results.append({
                                "rel_parts":    list(parts),
                                "analysis_dir": full,
                                "session_dir": current_dir,
                            })
                return

            # Still descending — list one level, recurse into subdirectories only
            try:
                entries = os.listdir(current_dir)
            except PermissionError:
                return
            for entry in entries:
                full = os.path.join(current_dir, entry)
                if os.path.isdir(full) and not entry.startswith("."):
                    _descend(full, parts + [entry], depth + 1)

        _descend(root_path, [], 0)
        return sorted(results, key=lambda item: item["analysis_dir"].casefold())

    # ── Fallback: full recursive walk (no depth rules available) ──────────
    for dirpath, dirnames, filenames in os.walk(root_path):
        if is_min1pipe and _contains_min1pipe_result(dirpath, filenames):
            rel = os.path.relpath(dirpath, root_path)
            rel_parts = [] if rel == "." else rel.replace("\\", "/").split("/")
            results.append({
                "rel_parts": rel_parts,
                "analysis_dir": dirpath,
                "session_dir": dirpath,
            })
            dirnames[:] = []
            continue
        # Reuse the directory listing already produced by os.walk. This avoids
        # a second network round-trip for every directory.
        matches = [name for name in dirnames if name.lower() == target]
        dirnames[:] = [
            name for name in dirnames
            if name.lower() != target and not name.startswith(".")
        ]
        for name in matches:
            rel = os.path.relpath(dirpath, root_path)
            rel_parts = [] if rel == "." else rel.replace("\\", "/").split("/")
            results.append({
                "rel_parts": rel_parts,
                "analysis_dir": os.path.join(dirpath, name),
                "session_dir": dirpath,
            })
    return sorted(results, key=lambda item: item["analysis_dir"].casefold())



# ──────────────────────────────────────────────────────────────────────────────
# Mapping helpers
# ──────────────────────────────────────────────────────────────────────────────

def apply_mapping(
    session: Dict[str, Any],
    depth_rules: List[Dict[str, Any]],
    global_values: Dict[str, Optional[str]],
) -> Optional[Dict[str, str]]:
    """
    Applies depth-labelling rules to the relative path parts of a session.

    *depth_rules* is a list (one element per depth level from root downward):
        {
            "label":        str   ("CohortName" | "MouseName" | "SessionType" |
                                   "SessionNumber" | "Ignore" | "Split")
            "split_regex":  str | None
            "split_fields": [str, str] | None   e.g. ["SessionType","SessionNumber"]
        }

    *global_values* fills any field not covered by a depth rule.

    Returns a dict { "CohortName": ..., "MouseName": ...,
                     "SessionType": ..., "SessionNumber": ... }
    or None if a required field is still missing.
    """
    parts = session["rel_parts"]
    result = {}

    for i, part in enumerate(parts):
        if i >= len(depth_rules):
            break
        rule = depth_rules[i]
        label = rule.get("label", "Ignore")

        if label == "Ignore":
            continue

        if label == "Split":
            pattern = rule.get("split_regex", "")
            fields  = rule.get("split_fields", [])
            if pattern and fields:
                m = re.match(pattern, part)
                if m and m.lastindex and m.lastindex >= len(fields):
                    for j, field in enumerate(fields):
                        result[field] = m.group(j + 1)
        else:
            result[label] = part

    # Fill in globals for any missing required field
    for field in REQUIRED_FIELDS:
        if field not in result:
            gval = global_values.get(field)
            if gval:
                result[field] = gval

    # Validate completeness
    for field in REQUIRED_FIELDS:
        if not result.get(field):
            return None  # incomplete mapping

    return result


def mapping_destination(mapping: Dict[str, str]) -> Tuple[str, str, str, str]:
    """Validate and return the complete HDF5 identity for a mapped session."""
    values = tuple(str(mapping[field]).strip() for field in REQUIRED_FIELDS)
    for field, value in zip(REQUIRED_FIELDS, values):
        if not value:
            raise ValueError(f"{field} cannot be empty.")
        if "/" in value or "\\" in value or any(ord(char) < 32 for char in value):
            raise ValueError(
                f"{field} contains a path separator or control character: {value!r}"
            )
    return values


def mapped_sessions(
    sessions: List[Dict[str, Any]],
    depth_rules: List[Dict[str, Any]],
    global_values: Dict[str, Optional[str]],
) -> Tuple[List[Tuple[Dict[str, Any], Dict[str, str]]], List[Dict[str, Any]]]:
    """Resolve destinations and report source folders mapped to the same key."""
    resolved = []
    destinations: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = {}
    for session in sessions:
        mapping = apply_mapping(session, depth_rules, global_values)
        if mapping is None:
            continue
        destination = mapping_destination(mapping)
        resolved.append((session, mapping))
        destinations.setdefault(destination, []).append(session)

    collisions = [
        {
            "destination": "/".join(destination),
            "sources": [item["analysis_dir"] for item in items],
        }
        for destination, items in destinations.items()
        if len(items) > 1
    ]
    return resolved, collisions


def _valid_hdf5_name(value: str) -> bool:
    return bool(value) and "/" not in value and "\\" not in value and not any(
        ord(char) < 32 for char in value
    )


def _table_stimulus_data(path: str) -> Tuple[Dict[str, np.ndarray], List[str]]:
    """Read a frame/stimulus CSV or XLSX table, dropping invalid rows."""
    suffix = os.path.splitext(path)[1].casefold()
    if suffix == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as stream:
            reader = csv.DictReader(stream)
            headers = list(reader.fieldnames or [])
            raw_rows = list(reader)
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(iterator, ())]
        raw_rows = [dict(zip(headers, row)) for row in iterator]
        workbook.close()
    else:
        raise ValueError("Per-session stimulus tables must be CSV or XLSX files.")
    normalized_headers = {str(value).strip().casefold() for value in headers}
    if not {"frame", "stimulus"}.issubset(normalized_headers):
        raise ValueError("Stimulus table must contain columns named 'frame' and 'stimulus'.")

    grouped: Dict[str, List[int]] = {}
    warnings = []
    for row_number, row in enumerate(raw_rows, start=2):
        normalized = {str(key).strip().casefold(): value for key, value in row.items()}
        raw_frame = normalized.get("frame")
        raw_stimulus = normalized.get("stimulus")
        try:
            numeric_frame = float(raw_frame)
            frame = int(numeric_frame)
            if not np.isfinite(numeric_frame) or numeric_frame != frame or frame < 1:
                raise ValueError
        except (TypeError, ValueError, OverflowError):
            warnings.append(f"row {row_number}: frame must be a positive 1-based integer")
            continue
        stimulus = "" if raw_stimulus is None else str(raw_stimulus).strip()
        if not _valid_hdf5_name(stimulus):
            warnings.append(f"row {row_number}: invalid stimulus name {stimulus!r}")
            continue
        grouped.setdefault(stimulus, []).append(frame)
    return {
        name: np.asarray(frames, dtype=np.int64)
        for name, frames in grouped.items()
    }, warnings


def _combined_stimulus_data(
    path: str,
    mapping: Dict[str, str],
) -> Dict[str, np.ndarray]:
    """Read one session's StimulusData group from a hierarchy-matched MAT file."""
    hierarchy = [ROOT_VAR] + [mapping[field] for field in REQUIRED_FIELDS] + ["StimulusData"]
    values: Dict[str, Any]
    try:
        import scipy.io

        loaded = scipy.io.loadmat(path, simplify_cells=True)
        current: Any = loaded
        for part in hierarchy:
            if not isinstance(current, dict) or part not in current:
                raise KeyError("/".join(hierarchy))
            current = current[part]
        if not isinstance(current, dict):
            raise ValueError(f"{'/'.join(hierarchy)} is not a structure.")
        values = current
    except (NotImplementedError, ValueError, OSError):
        with h5py.File(path, "r") as handle:
            h5_path = "/".join(hierarchy)
            if h5_path not in handle or not isinstance(handle[h5_path], h5py.Group):
                raise KeyError(h5_path)
            values = {
                name: np.asarray(obj[()])
                for name, obj in handle[h5_path].items()
                if isinstance(obj, h5py.Dataset)
            }

    output = {}
    for name, raw_frames in values.items():
        if not _valid_hdf5_name(str(name)):
            raise ValueError(f"Invalid stimulus name {name!r} in {'/'.join(hierarchy)}.")
        frames = np.asarray(raw_frames).reshape(-1)
        if not np.issubdtype(frames.dtype, np.number):
            raise ValueError(f"Stimulus {name!r} must contain numeric frame numbers.")
        numeric = frames.astype(np.float64)
        if np.any(~np.isfinite(numeric)) or np.any(numeric < 1) or np.any(numeric != np.floor(numeric)):
            raise ValueError(f"Stimulus {name!r} contains invalid 1-based frame numbers.")
        output[str(name)] = numeric.astype(np.int64)
    return output


def _session_stimulus_data(
    session: Dict[str, Any],
    mapping: Dict[str, str],
    analysis_type: str,
    mode: str,
    table_pattern: Optional[str],
    combined_path: Optional[str],
) -> Tuple[Optional[Dict[str, np.ndarray]], List[str]]:
    if mode in {"", "none", None}:
        return None, []
    if mode == "combined":
        if not combined_path:
            raise ValueError("A combined stimulus MAT/HDF5 file is required.")
        return _combined_stimulus_data(combined_path, mapping), []
    if mode != "table":
        raise ValueError("stimulus_mode must be 'none', 'table', or 'combined'.")
    if not table_pattern:
        raise ValueError("A relative per-session stimulus table pattern is required.")
    formatted = table_pattern.format(**mapping)
    if os.path.isabs(formatted):
        raise ValueError("Per-session stimulus table patterns must be relative paths.")
    session_dir = session.get("session_dir")
    if not session_dir:
        session_dir = (
            session["analysis_dir"]
            if analysis_type == "min1pipe"
            else os.path.dirname(session["analysis_dir"])
        )
    matches = sorted(glob.glob(os.path.join(session_dir, formatted)))
    if not matches:
        return None, [f"stimulus table not found: {formatted}"]
    data, warnings = _table_stimulus_data(matches[0])
    if len(matches) > 1:
        warnings.insert(0, f"multiple stimulus tables matched; used {os.path.basename(matches[0])}")
    return data, warnings


# ──────────────────────────────────────────────────────────────────────────────
# HDF5 writer helpers
# ──────────────────────────────────────────────────────────────────────────────

def _dataset_chunks(name: str, shape: Tuple[int, ...]) -> Optional[Tuple[int, ...]]:
    if not shape or any(size == 0 for size in shape):
        return None
    if name == "SpatialFootprints":
        return (1, min(shape[1], 256), min(shape[2], 256))
    if name in {"TemporalFootprints", "DeconvolvedEvents", "DeltaFOverF"}:
        return (min(shape[0], 32), min(shape[1], 2048))
    if name == "MaxProjection":
        return tuple(min(size, 256) for size in shape)
    return None


def _write_dataset(
    grp: h5py.Group,
    name: str,
    data: np.ndarray,
    h5_path: str,
    dtype: np.dtype,
    compression: Optional[str],
    compression_level: int,
):
    """Creates or overwrites a dataset inside *grp* with MATLAB-compatible attrs."""
    if name in grp:
        del grp[name]
    options: Dict[str, Any] = {"chunks": _dataset_chunks(name, data.shape)}
    if compression:
        options.update(compression=compression, shuffle=True)
        if compression == "gzip":
            options["compression_opts"] = compression_level
    ds = grp.create_dataset(name, data=data, dtype=dtype, **options)
    if np.issubdtype(np.dtype(dtype), np.integer):
        matlab_class = b"int8" if np.dtype(dtype).itemsize == 1 else b"int64"
    else:
        matlab_class = b"single" if np.dtype(dtype) == np.dtype("float32") else b"double"
    ds.attrs["MATLAB_class"] = matlab_class
    ds.attrs["H5PATH"] = h5_path.encode("utf-8")


def _ensure_group(f: h5py.File, path: str) -> h5py.Group:
    """Creates intermediate groups if they don't exist."""
    if path not in f:
        f.require_group(path)
    return f[path]


def _validated_arrays(data: Dict[str, Any]) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """Validate the loader contract before altering an existing session."""
    missing = [name for name in ("spatial", "temporal", "max_proj") if name not in data]
    if missing:
        raise ValueError(f"Loader result is missing: {', '.join(missing)}")
    spatial = np.asarray(data["spatial"])
    temporal = np.asarray(data["temporal"])
    max_projection = np.asarray(data["max_proj"])
    if spatial.ndim != 3:
        raise ValueError(f"Spatial footprints must be (N,H,W), got {spatial.shape}.")
    if temporal.ndim != 2:
        raise ValueError(f"Temporal footprints must be (N,T), got {temporal.shape}.")
    if max_projection.ndim != 2:
        raise ValueError(f"Max projection must be (H,W), got {max_projection.shape}.")
    if spatial.shape[0] != temporal.shape[0]:
        raise ValueError(
            f"Cell count mismatch: spatial={spatial.shape[0]}, temporal={temporal.shape[0]}."
        )
    if tuple(spatial.shape[1:]) != tuple(max_projection.shape):
        raise ValueError(
            f"FOV mismatch: spatial={spatial.shape[1:]}, MIP={max_projection.shape}."
        )
    if spatial.shape[0] == 0 or temporal.shape[1] == 0:
        raise ValueError("A session must contain at least one cell and one frame.")
    if not all(
        np.issubdtype(array.dtype, np.number)
        for array in (spatial, temporal, max_projection)
    ):
        raise ValueError("All calcium datasets must be numeric.")
    return spatial, temporal, max_projection


def _publish_session(
    f: h5py.File,
    session_parent: h5py.Group,
    session_parent_path: str,
    session_name: str,
    data: Dict[str, Any],
    spatial: np.ndarray,
    temporal: np.ndarray,
    max_projection: np.ndarray,
    source_relative_path: str,
    analysis_type: str,
    precision: str,
    compression: Optional[str],
    compression_level: int,
) -> None:
    """Stage a complete session and swap it in without exposing partial data."""
    session_path = f"{session_parent_path}/{session_name}"
    temp_name = f".__building_{session_name}_{uuid4().hex}"
    backup_name = f".__previous_{session_name}_{uuid4().hex}"
    temp_path = f"{session_parent_path}/{temp_name}"
    backup_path = f"{session_parent_path}/{backup_name}"
    cal_path = f"{session_path}/CalciumData"
    backup_created = False

    try:
        temp_session = session_parent.create_group(temp_name)
        temp_session.attrs["SourceRelativePath"] = source_relative_path
        temp_session.attrs["AnalysisType"] = analysis_type
        temp_session.attrs["MaxProjectionSource"] = data.get(
            "max_projection_source", "unspecified"
        )
        temp_session.attrs["CellCount"] = int(spatial.shape[0])
        temp_session.attrs["FrameCount"] = int(temporal.shape[1])
        cal_grp = temp_session.create_group("CalciumData")
        storage_dtype = np.dtype(precision)

        _write_dataset(
            cal_grp,
            "SpatialFootprints",
            np.transpose(spatial, (0, 2, 1)),
            f"/{cal_path}",
            storage_dtype,
            compression,
            compression_level,
        )
        _write_dataset(
            cal_grp,
            "TemporalFootprints",
            temporal,
            f"/{cal_path}",
            storage_dtype,
            compression,
            compression_level,
        )
        _write_dataset(
            cal_grp,
            "MaxProjection",
            max_projection.T,
            f"/{cal_path}",
            storage_dtype,
            compression,
            compression_level,
        )
        for source_name, dataset_name in (
            ("deconvolved_events", "DeconvolvedEvents"),
            ("delta_f_over_f", "DeltaFOverF"),
        ):
            if source_name not in data:
                continue
            optional = np.asarray(data[source_name])
            if optional.ndim != 2 or optional.shape != temporal.shape:
                raise ValueError(
                    f"{dataset_name} must match TemporalFootprints shape "
                    f"{temporal.shape}; got {optional.shape}."
                )
            _write_dataset(
                cal_grp,
                dataset_name,
                optional,
                f"/{cal_path}",
                storage_dtype,
                compression,
                compression_level,
            )

        quality_grp = cal_grp.create_group("CellQuality")
        common_quality = cell_quality_metrics(spatial, temporal)
        for name, values in common_quality.items():
            _write_dataset(
                quality_grp,
                name,
                values,
                f"/{cal_path}/CellQuality",
                storage_dtype,
                compression,
                compression_level,
            )
        source_quality = data.get("source_quality") or {}
        valid_source_quality = {
            name: np.asarray(values)
            for name, values in source_quality.items()
            if name in {"Accepted", "TemporalSNR", "SpatialCorrelation", "ClassifierScore"}
            and np.asarray(values).reshape(-1).size == spatial.shape[0]
        }
        if valid_source_quality:
            source_grp = quality_grp.create_group("Source")
            for name, values in valid_source_quality.items():
                values = values.reshape(-1)
                dtype = np.dtype("int8") if name == "Accepted" else storage_dtype
                _write_dataset(
                    source_grp,
                    name,
                    values,
                    f"/{cal_path}/CellQuality/Source",
                    dtype,
                    compression,
                    compression_level,
                )

        stimulus_data = data.get("stimulus_data")
        if stimulus_data is not None:
            stimulus_grp = temp_session.create_group("StimulusData")
            for name, frames in stimulus_data.items():
                _write_dataset(
                    stimulus_grp,
                    name,
                    np.asarray(frames, dtype=np.int64).reshape(-1),
                    f"/{session_path}/StimulusData",
                    np.dtype("int64"),
                    compression,
                    compression_level,
                )
        f.flush()

        if session_name in session_parent:
            f.move(session_path, backup_path)
            backup_created = True
        f.move(temp_path, session_path)
        if backup_created:
            del session_parent[backup_name]
        f.flush()
    except Exception:
        if temp_name in session_parent:
            del session_parent[temp_name]
        if backup_created and session_name not in session_parent and backup_name in session_parent:
            f.move(backup_path, session_path)
        f.flush()
        raise


# ──────────────────────────────────────────────────────────────────────────────
# Build
# ──────────────────────────────────────────────────────────────────────────────

def build(
    output_path: str,
    root_path: str,
    analysis_type: str,
    depth_rules: List[Dict[str, Any]],
    global_values: Dict[str, Optional[str]],
    progress_callback: Callable[[str], None] = print,
    precision: str = "float64",
    compression: Optional[str] = "gzip",
    compression_level: int = 4,
    append_policy: str = "replace",
    stimulus_mode: str = "none",
    stimulus_table_pattern: Optional[str] = None,
    stimulus_combined_path: Optional[str] = None,
    analysis_pattern: Optional[str] = None,
) -> int:
    """
    Discovers sessions under *root_path*, maps their paths, loads data, and
    writes into an HDF5 file at *output_path*.

    Returns the number of sessions successfully written.
    """
    if precision not in {"float32", "float64"}:
        raise ValueError("precision must be 'float32' or 'float64'.")
    if compression in {"", "none", "None"}:
        compression = None
    if compression not in {None, "gzip", "lzf"}:
        raise ValueError("compression must be 'gzip', 'lzf', or 'none'.")
    if not 0 <= compression_level <= 9:
        raise ValueError("compression_level must be between 0 and 9.")
    if append_policy not in {"replace", "skip", "error"}:
        raise ValueError("append_policy must be 'replace', 'skip', or 'error'.")

    loader = get_loader(analysis_type)
    sessions = discover_sessions(
        root_path,
        analysis_type,
        depth_rules=depth_rules,
        analysis_pattern=analysis_pattern,
    )

    if not sessions:
        progress_callback(f"⚠  No '{analysis_type}' folders found under: {root_path}")
        return 0

    progress_callback(f"🔍  Found {len(sessions)} '{analysis_type}' folder(s) to process…")

    _, collisions = mapped_sessions(sessions, depth_rules, global_values)
    if collisions:
        details = "; ".join(
            f"{item['destination']} <- {len(item['sources'])} source folders"
            for item in collisions
        )
        raise ValueError(f"Duplicate destination mappings detected: {details}")

    written = 0

    with h5py.File(output_path, "a") as f:
        # Ensure root variable group exists
        root_grp = _ensure_group(f, ROOT_VAR)
        root_grp.attrs["SchemaName"] = "Calcium Imaging Database"
        root_grp.attrs["SchemaVersion"] = "1.2"
        root_grp.attrs["Builder"] = "db_builder_py"
        root_grp.attrs["LastBuildUTC"] = datetime.now(timezone.utc).isoformat()

        for sess in sessions:
            # Apply user mapping
            mapping = apply_mapping(sess, depth_rules, global_values)

            if mapping is None:
                label_str = "/".join(sess["rel_parts"]) if sess["rel_parts"] else "(root)"
                progress_callback(
                    f"⚠  Skipped '{label_str}': could not resolve all required fields "
                    f"(CohortName, MouseName, SessionType, SessionNumber)."
                )
                continue

            cohort, mouse, stype, snum = mapping_destination(mapping)

            label_str = f"{cohort}/{mouse}/{stype}/{snum}"

            # Load data
            try:
                data = loader.load(sess["analysis_dir"])
            except DataNotFoundError as exc:
                progress_callback(f"⚠  {label_str}: {exc}")
                continue
            except Exception as exc:
                progress_callback(f"✗  {label_str}: Unexpected error — {exc}")
                continue

            # Build HDF5 path
            try:
                sf, tf, mp = _validated_arrays(data)
            except ValueError as exc:
                progress_callback(f"Invalid {label_str}: {exc}")
                continue

            try:
                stimulus_data, stimulus_warnings = _session_stimulus_data(
                    sess,
                    mapping,
                    analysis_type,
                    stimulus_mode,
                    stimulus_table_pattern,
                    stimulus_combined_path,
                )
                if stimulus_data is not None:
                    data["stimulus_data"] = stimulus_data
                for warning in stimulus_warnings:
                    progress_callback(f"Warning {label_str}: {warning}.")
            except (KeyError, ValueError, OSError) as exc:
                progress_callback(f"Warning {label_str}: stimulus data was not imported ({exc}).")

            session_parent_path = f"{ROOT_VAR}/{cohort}/{mouse}/{stype}"
            session_path = f"{session_parent_path}/{snum}"
            session_parent = _ensure_group(f, session_parent_path)
            if snum in session_parent and append_policy == "skip":
                progress_callback(f"Skipped {label_str}: session already exists.")
                continue
            if snum in session_parent and append_policy == "error":
                raise ValueError(f"Session already exists in output: {label_str}")

            _publish_session(
                f,
                session_parent,
                session_parent_path,
                snum,
                data,
                sf,
                tf,
                mp,
                os.path.relpath(sess["analysis_dir"], root_path).replace("\\", "/"),
                analysis_type,
                precision,
                compression,
                compression_level,
            )

            # SpatialFootprints: store as (N, W, H) = transpose of (N, H, W)

            # TemporalFootprints: (N, T) — no transposition needed

            # MaxProjection: store as (W, H) = transpose of (H, W)

            progress_callback(
                f"✓  {label_str}  "
                f"[{sf.shape[0]} cells, {tf.shape[1]} frames]"
            )
            written += 1

    progress_callback(
        f"\n{'='*50}\n"
        f"Build complete — {written}/{len(sessions)} sessions written.\n"
        f"Output: {output_path}"
    )
    return written
